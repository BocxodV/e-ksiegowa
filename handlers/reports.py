# handlers/reports.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Router, types, F, Bot
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile

from database import get_user_profile, get_available_months, get_work_logs_for_month
from texts import TRANSLATIONS

router = Router()

@router.message(F.text.in_(["📊 Mój raport", "📊 Мій звіт", "📊 Мой отчет", "📊 My Report"]))
async def ask_report_month(message: types.Message):
    user_id = message.from_user.id
    profile = await get_user_profile(user_id)
    
    user_lang = profile.get("lang", "RUS")
    t = TRANSLATIONS.get(user_lang, TRANSLATIONS["RUS"])

    months = await get_available_months(user_id)

    if not months:
        await message.answer(t["empty_db"])
        return

    all_buttons = [InlineKeyboardButton(text=f"📅 {m[0]}", callback_data=f"report_{m[0]}") for m in months]
    rows_of_3 = [all_buttons[i:i+3] for i in range(0, len(all_buttons), 3)]
    keyboard = InlineKeyboardMarkup(inline_keyboard=rows_of_3)
    await message.answer(t["choose_month"], reply_markup=keyboard)

@router.callback_query(F.data.startswith("report_"))
async def generate_excel_report(callback: types.CallbackQuery = None, target_user_id=None, target_month=None, bot: Bot = None):
    # Interceptor: Notify the user that processing has started
    if callback:
        user_id = callback.from_user.id
        selected_month = callback.data.split("_")[1]
        await callback.message.delete()
        # Optional placeholder for custom log message
        await callback.bot.send_chat_action(chat_id=callback.message.chat.id, action="upload_document")
    else:
        user_id = target_user_id
        selected_month = target_month

    profile = await get_user_profile(user_id)
    
    user_lang = profile.get("lang", "RUS")
    t = TRANSLATIONS.get(user_lang, TRANSLATIONS["RUS"])

    rows = await get_work_logs_for_month(user_id, selected_month)

    if not rows:
        if callback:
            await callback.message.answer(t["empty_db"])
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Raport {selected_month}"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")

    urlop_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    l4_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    headers = t["excel_headers"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned

    total_net, total_gross, total_bonuses, total_hours, total_drive_hours = 0.0, 0.0, 0.0, 0.0, 0.0

    for row in rows:
        is_trip_text = "✅" if row[10] == 1 else ""

        excel_row = [
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9],
            is_trip_text, row[11], row[12], row[13]
        ]
        ws.append(excel_row)

        current_row_num = ws.max_row
        status = row[2]
        fill_to_apply = urlop_fill if status == "Urlop" else l4_fill if status == "L4" else None

        for col_num, cell in enumerate(ws[current_row_num], start=1):
            if fill_to_apply:
                cell.fill = fill_to_apply
            if col_num in [4, 5, 6]:
                cell.alignment = left_aligned
            else:
                cell.alignment = center_aligned

        total_hours += float(row[6] or 0)
        total_drive_hours += float(row[7] or 0)
        total_bonuses += float(row[11] or 0)
        total_gross += float(row[12] or 0)
        total_net += float(row[13] or 0)

    ws.append([]) 

    tax_coeff = profile.get("tax_coeff", 0.72)
    card_money = total_gross * tax_coeff
    envelope_money = total_net - card_money
    if envelope_money < 0: envelope_money = 0

    total_row = [
        "", "", "", "", t["total_month"], "", round(total_hours, 1), round(total_drive_hours, 1), "", "", "",
        round(total_bonuses, 2), round(total_gross, 2), round(total_net, 2)
    ]
    ws.append(total_row)

    ws.append([])
    ws.append(["", "", "", "", "", "💳 НА КАРТУ (Официально):", "", "", "", "", "", "", round(card_money, 2), "zł"])
    ws.append(["", "", "", "", "", "⚖️ Audyt wyrównania:", "", "", "", "", "", "", round(envelope_money, 2), "zł"])

    for row_idx in range(ws.max_row - 2, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.alignment = center_aligned

    ws.append([])
    last_row = ws.max_row + 1
    ws.merge_cells(start_row=last_row, start_column=11, end_row=last_row, end_column=14)
    sig_cell = ws.cell(row=last_row, column=11)
    sig_cell.value = "© Created by bocxodv"
    sig_cell.font = Font(italic=True, size=9, color="A6A6A6")
    sig_cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[col[0].column_letter].width = (max_length * 1.25) + 3

    # Store the spreadsheet directly in RAM
    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0) # Rewind the buffer pointer to the beginning
    
    file_name = f"Zarobki_{selected_month}.xlsx"
    document = BufferedInputFile(file_buffer.read(), filename=file_name)

    caption_text = t["excel_caption"].format(
        month=selected_month,
        hours=round(total_hours, 1),
        net=round(total_net, 2)
    )

    try:
        if callback:
            await callback.message.answer_document(document, caption=caption_text, parse_mode="Markdown")
        elif bot:
            await bot.send_document(user_id, document, caption=caption_text, parse_mode="Markdown")
            
    except Exception as e:
        print(f"Error sending report: {e}")
        if callback:
            await callback.message.answer("⚠️ Ошибка обработки данных.")
    # Garbage collection will automatically clean up the BytesIO buffer

async def generate_boss_excel_report(target_user_id, target_month, bot: Bot):
    profile = await get_user_profile(target_user_id)
    user_lang = profile.get("lang", "RUS")
    t = TRANSLATIONS.get(user_lang, TRANSLATIONS["RUS"])

    rows = await get_work_logs_for_month(target_user_id, target_month)

    if not rows:
        await bot.send_message(target_user_id, t["empty_db"])
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Logistyka {target_month}"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")

    urlop_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    l4_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    headers = t.get("excel_boss_headers", ["Data", "Dzień", "Status", "Obiekt", "Auto", "Trasa", "Suma h", "Jazda", "50%", "100%"])
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned

    total_hours, total_drive_hours = 0.0, 0.0
    total_50, total_100 = 0.0, 0.0

    for row in rows:
        excel_row = [
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]
        ]
        ws.append(excel_row)

        current_row_num = ws.max_row
        status = row[2]
        fill_to_apply = urlop_fill if status == "Urlop" else l4_fill if status == "L4" else None

        for col_num, cell in enumerate(ws[current_row_num], start=1):
            if fill_to_apply:
                cell.fill = fill_to_apply
            if col_num in [4, 5, 6]:
                cell.alignment = left_aligned
            else:
                cell.alignment = center_aligned

        total_hours += float(row[6] or 0)
        total_drive_hours += float(row[7] or 0)
        total_50 += float(row[8] or 0)
        total_100 += float(row[9] or 0)

    ws.append([]) 

    total_row = [
        "", "", "", "", t["total_month"], "", round(total_hours, 1), round(total_drive_hours, 1), round(total_50, 1), round(total_100, 1)
    ]
    ws.append(total_row)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = center_aligned

    ws.append([])
    last_row = ws.max_row + 1
    ws.merge_cells(start_row=last_row, start_column=7, end_row=last_row, end_column=10)
    sig_cell = ws.cell(row=last_row, column=7)
    sig_cell.value = "© Created by bocxodv"
    sig_cell.font = Font(italic=True, size=9, color="A6A6A6")
    sig_cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[col[0].column_letter].width = (max_length * 1.25) + 3

    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    file_name = f"Logistyka_{target_month}.xlsx"
    document = BufferedInputFile(file_buffer.read(), filename=file_name)

    caption_text = t.get("excel_boss_caption", "📊 Raport operacyjny za {month}\n🕒 Przepracowano: **{hours} h.**").format(
        month=target_month,
        hours=round(total_hours, 1)
    )

    try:
        await bot.send_document(target_user_id, document, caption=caption_text, parse_mode="Markdown")
    except Exception as e:
        print(f"Failed to send boss report: {e}")

async def generate_pure_logistics_report(target_user_id, target_month, bot: Bot):
    profile = await get_user_profile(target_user_id)
    user_lang = profile.get("lang", "RUS")
    t = TRANSLATIONS.get(user_lang, TRANSLATIONS["RUS"])

    rows = await get_work_logs_for_month(target_user_id, target_month)

    if not rows:
        await bot.send_message(target_user_id, t["empty_db"])
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transport {target_month}"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center", vertical="center")

    headers = t.get("excel_pure_logistics_headers", ["Data", "Dzień", "Trasa", "Auto", "Godziny jazdy"])
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned

    total_drive_hours = 0.0

    for row in rows:
        # row: 0=Date, 1=Day, 2=Status, 3=Object, 4=Car, 5=Route, 6=Work, 7=Driving
        driving = float(row[7] or 0)
        # Skip rows where driving is 0 or it's not a work day
        if row[2] != "Work" or driving <= 0:
            continue

        excel_row = [
            row[0], row[1], row[5], row[4], row[7]
        ]
        ws.append(excel_row)
        
        for cell in ws[ws.max_row]:
            cell.alignment = center_aligned

        total_drive_hours += driving

    ws.append([]) 

    total_row = [
        "", "", "", t["total_month"], round(total_drive_hours, 1)
    ]
    ws.append(total_row)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = center_aligned

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[col[0].column_letter].width = (max_length * 1.25) + 3

    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    file_name = f"Transport_{target_month}.xlsx"
    document = BufferedInputFile(file_buffer.read(), filename=file_name)

    caption_text = t.get("excel_pure_logistics_caption", "🚚 Raport dla logistyka za {month}\n🚗 Całkowity czas za kierownicą: **{hours} h.**").format(
        month=target_month,
        hours=round(total_drive_hours, 1)
    )

    try:
        await bot.send_document(target_user_id, document, caption=caption_text, parse_mode="Markdown")
    except Exception as e:
        print(f"Failed to send pure logistics report: {e}")